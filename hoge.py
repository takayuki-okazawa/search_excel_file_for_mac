#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys, codecs, os
sys.stdout = codecs.getwriter('utf_8')(sys.stdout)

from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.styles import PatternFill

keyword = "売上カレンダー"

def fild_all_files(directory):
    for root, dirs, files in os.walk(directory):
        yield root
        for file in files:
            yield os.path.join(root, file)

for name in fild_all_files(u"./"):
    #print(name)
    if -1 < name.find(".XLSX"):
        wb = None
        try:
            wb = load_workbook(filename=name)
        except:
            continue
        ws = wb.active
        x = ws.rows

        for u in x:
          for v in u:
            if v.value != None:
                if isinstance(v.value, unicode):
                    str2 = v.value.encode('utf-8')
                    if -1 < str2.find(keyword):
                        print(name)
                        print(name + u" 該当セル : " + v.value)
